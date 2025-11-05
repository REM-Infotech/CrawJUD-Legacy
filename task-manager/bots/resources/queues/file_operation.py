"""Operações de planilhas."""

from contextlib import suppress
from datetime import datetime
from queue import Empty, Queue
from threading import Thread
from typing import TYPE_CHECKING, NoReturn
from zoneinfo import ZoneInfo

import openpyxl
from pandas import DataFrame, ExcelWriter, concat, read_excel

from app.types import Dict

DATASAVE = []

if TYPE_CHECKING:
    from bots.head import CrawJUD


class SaveSuccess:
    """Controle da Queue de salvamento de sucessos."""

    bot: CrawJUD

    def __init__(self, bot: CrawJUD) -> None:
        """Instancia da queue de salvamento de erros."""
        self.bot = bot

        self.queue_save = Queue()
        self.thead_save = Thread(target=self.save_success, daemon=True)
        self.thead_save.start()

    def __call__(self, work_sheet: str, data_save: str) -> None:
        self.queue_save.put_nowait({
            "work_sheet": work_sheet,
            "data_save": DataFrame(data_save).to_dict(orient="records"),
        })

    def save_success(self) -> NoReturn:
        tz = ZoneInfo("America/Sao_Paulo")
        now = datetime.now(tz=tz).strftime("%d-%m-%Y %H-%M-%S")
        nome_arquivo = f"Sucessos - PID {self.bot.pid} - {now}.xlsx"
        arquivo_sucesso = self.bot.output_dir_path.joinpath(nome_arquivo)
        while True:
            data: Dict | None = None

            with suppress(Empty):
                data = self.queue_save.get_nowait()

            if data:
                with suppress(Exception):
                    df = DataFrame(data["data_save"])
                    if arquivo_sucesso.exists():
                        wb = openpyxl.load_workbook(str(arquivo_sucesso))
                        writer = ExcelWriter(
                            arquivo_sucesso,
                            engine="openpyxl",
                            mode="a",
                            if_sheet_exists="replace",
                        )
                        # Verifica se a worksheet já existe
                        if data["work_sheet"] in wb.sheetnames:
                            df_xlsx = read_excel(
                                arquivo_sucesso,
                                engine="openpyxl",
                                sheet_name=data["work_sheet"],
                            )
                            df = concat([
                                DataFrame(df_xlsx.to_dict(orient="records")),
                                DataFrame(data["data_save"]),
                            ])
                    else:
                        writer = ExcelWriter(arquivo_sucesso, engine="openpyxl")

                    df.to_excel(
                        excel_writer=writer,
                        sheet_name=data["work_sheet"],
                        index=False,
                    )
                    writer.close()


class SaveError:
    """Controle da Queue de salvamento de erros."""

    def __init__(self, bot: CrawJUD) -> None:
        """Instancia da queue de salvamento de erros."""
        self.bot = bot
        self.queue_save = Queue()
        self.thead_save = Thread(target=self.save_error, daemon=True)
        self.thead_save.start()

    def __call__(self, work_sheet: str, data_save: list[Dict]) -> None:
        self.queue_save.put_nowait({
            "work_sheet": work_sheet,
            "data_save": DataFrame(data_save).to_dict(orient="records"),
        })

    def save_error(self) -> NoReturn:
        tz = ZoneInfo("America/Sao_Paulo")
        now = datetime.now(tz=tz).strftime("%d-%m-%Y %H-%M-%S")
        nome_arquivo = f"Erros - PID {self.bot.pid} - {now}.xlsx"
        arquivo_sucesso = self.bot.output_dir_path.joinpath(nome_arquivo)
        while True:
            data: Dict | None = None

            with suppress(Empty):
                data = self.queue_save.get_nowait()

            if data:
                with suppress(Exception):
                    df = DataFrame(data["data_save"])
                    if arquivo_sucesso.exists():
                        wb = openpyxl.load_workbook(str(arquivo_sucesso))
                        writer = ExcelWriter(
                            arquivo_sucesso,
                            engine="openpyxl",
                            mode="a",
                            if_sheet_exists="replace",
                        )
                        # Verifica se a worksheet já existe
                        if data["work_sheet"] in wb.sheetnames:
                            df_xlsx = read_excel(
                                arquivo_sucesso,
                                engine="openpyxl",
                                sheet_name=data["work_sheet"],
                            )
                            df = concat([
                                DataFrame(df_xlsx.to_dict(orient="records")),
                                DataFrame(data["data_save"]),
                            ])
                    else:
                        writer = ExcelWriter(arquivo_sucesso, engine="openpyxl")

                    df.to_excel(
                        excel_writer=writer,
                        sheet_name=data["work_sheet"],
                        index=False,
                    )
                    writer.close()
