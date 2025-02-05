"""Module: MakeXlsx.

This module provides functionality to create and customize Excel files.

using openpyxl, based on the attributes of the CrawJUD framework.
"""

from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill

from ...core import CrawJUD
from .appends import listas


class MakeXlsx(CrawJUD):
    """Creates an Excel file with customized headers and styles."""

    def __init__(self) -> None:
        """Initialize the MakeXlsx class with default settings."""

    def make_output(self, type_xlsx: str, path_template: str) -> List[str]:
        """Build and save a new Excel file at the given path with various headers.

        Args:
            type_xlsx (str): A string defining the type of Excel template.
            path_template (str): The path to save the generated Excel file.

        Returns:
            List[str]: A list of the headers used in the created Excel file.

        """
        lista_colunas: List[str] = getattr(listas(), f"{self.typebot}_{type_xlsx}", getattr(listas(), type_xlsx, None))
        # Criar um novo workbook e uma planilha
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("Resultados", 0)
        sheet = workbook.active

        # Cabeçalhos iniciais
        cabecalhos = ["NUMERO_PROCESSO"]
        list_to_append = []

        list_to_append.extend(lista_colunas)
        cabecalhos.extend(list_to_append)

        # Definir estilo
        my_red = openpyxl.styles.colors.Color(rgb="A6A6A6")
        my_fill = PatternFill(patternType="solid", fgColor=my_red)
        bold_font = Font(name="Times New Roman", italic=True)

        # Escrever os cabeçalhos na primeira linha da planilha e aplicar estilo
        for pos, coluna in enumerate(cabecalhos):
            item = sheet.cell(row=1, column=pos + 1, value=coluna.upper())
            item.font = bold_font
            item.fill = my_fill

        # Ajustar a largura das colunas
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    raise e
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

        # Salvar o workbook no caminho especificado
        workbook.save(path_template)

        return cabecalhos
